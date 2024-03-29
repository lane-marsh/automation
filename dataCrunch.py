import csv
import openpyxl


class ObjectifyXL(object):
    """
    object for parsing table in an excel file
    """

    def __init__(self, path, head_row=1, sheet=''):
        """
        input parameter:    path
        description:        relative path from the python file running to the target xlsx file
        input parameter:    head_row
        description:        row in the worksheet with column headers
                            defaults to the first row
        input parameter:    sheet
        description:        name of the sheet with the table
                            defaults to the active sheet when the file is opened
        """

        self.data = {}
        self.headers = []
        self.total_entries = 0
        self.path = path
        self.offset = head_row

        wb = openpyxl.load_workbook(self.path)
        self.shape_data(wb, head_row, sheet)
        wb.close()

    def shape_data(self, wb, head_row, sheet):
        """
        take workbook and optionally sheet name
        populate the data dictionary with contents of the workbook sheet
        """

        row_counter = 0
        col_counter = 1

        if sheet == '':
            ws = wb.active
        else:
            ws = wb.get_sheet_by_name(sheet)

        max_rows = ws.max_row
        self.total_entries = max_rows - head_row

        while ws.cell(row=head_row, column=col_counter).value is not None:
            title = ws.cell(row=head_row, column=col_counter).value
            self.headers.append(title)
            self.data[title] = {}
            col_counter += 1

        for row in range(head_row+1, max_rows+1):
            row_counter += 1
            col_counter = 1
            for title in self.headers:
                entry = ws.cell(row=row, column=col_counter).value
                self.data[title][row_counter] = entry
                col_counter += 1

    def get_by_field(self, filt, return_fields=None):
        """
        input parameter:    filter
        description:        dictionary that matches field names to a value that we filter by
        optional input:     return_fields
        description:        use if only specific fields are wanted to be returned.
                            defaults to returning all fields.
        """

        if return_fields is None:
            return_fields = self.headers

        results = []
        keys = set()

        # search for any key that matches a filter and add it to the keys set
        for header, sets in self.data.items():
            for key, value in sets.items():
                if header in filt:
                    if filt[header] == value:
                        keys.add(key)

        for key in keys:
            row = []
            for header in return_fields:
                row.append(self.data[header][key])
            results.append(row)

        return results

    def write_to_file(self, in_dict, sheet_name=None, keys=None):
        """
        input parameter:    in_dict
        description:        dictionary where keys match a field in the data and their
                            corresponding values are what will populate a new line
                            option to input multiple fields with one call by populating
                            dictionary with a list.
        assumption:         if provided key does not match a field, it will be ignored.
        """

        file = openpyxl.load_workbook(self.path, read_only=False)

        if sheet_name:
            table = file['Sheet1']
        else:
            table = file.active

        if not keys:
            # add to end of data file
            start_key = len(self.data[self.headers[0]]) + 1
            for header, data_list in in_dict.items():
                entry_row = start_key + self.offset
                entry_col = self.headers.index(header) + 1
                for value in data_list:
                    table.cell(row=entry_row, column=entry_col).value = value
                    entry_row += 1
        else:
            # keys provided, overwrite existing data fields
            for header, data_list in in_dict.items():
                entry_col = self.headers.index(header) + 1
                for index, value in enumerate(data_list):
                    entry_row = keys[index] + self.offset
                    table.cell(row=entry_row, column=entry_col).value = value

        file.save(self.path)
        file.close()
        return True


class ObjectifyCSV(object):
    """
    object for parsing a csv file
    """

    def __init__(self, path):
        """
        input parameter:    path
        description:        relative path from the python file running to the target csv file
        """

        self.data = {}
        self.headers = []
        self.total_entries = 0
        self.path = path
        self.file_str = ''

        with open(path, 'r') as csv_file:
            self.file_str = csv_file.read()
        csv_file.close()

        with open(path) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            line_count = 0
            for row in csv_reader:
                if line_count == 0:
                    for header in row:
                        self.data[header] = {}
                        self.headers.append(header)
                else:
                    self.total_entries += 1
                    col_count = 0
                    for entry in row:
                        header = self.headers[col_count]
                        self.data[header][line_count] = entry
                        col_count += 1
                line_count += 1

    def get_by_field(self, filter, return_fields=None):
        """
        input parameter:    filter
        description:        dictionary that matches field names to a value that we filter by
        optional input:     return_fields
        description:        use if only specific fields are wanted to be returned.
                            defaults to returning all fields.
        """

        if return_fields is None:
            return_fields = self.headers

        results = []
        keys = set()

        # search for any key that matches a filter and add it to the keys set
        for header, sets in self.data.items():
            for key, value in sets.items():
                if header in filter:
                    if filter[header] == value:
                        keys.add(key)

        for key in keys:
            row = []
            for header in return_fields:
                row.append(self.data[header][key])
            results.append(row)

        return results

    def write_to_file(self, in_dict):
        """
        input parameter:    in_dict
        description:        dictionary where keys match a field in the data and their
                            corresponding values are what will populate a new line
                            option to input multiple fields with one call by populating
                            dictionary with a list.
        assumption:         if provided key does not match a field, it will be ignored.
        """

        def arr_to_entry(in_arr):
            result = ''
            for value in in_arr:
                if value:
                    result += value
                result += ','
            return result[:-1] + '\n'

        multiple_fields = None
        entries = 1

        for key, value in in_dict.items():
            if type(value) == list:
                if multiple_fields or multiple_fields is None:
                    multiple_fields = True
                    entries = len(value)
                else:
                    print('error, incompatible data entry', value, 'expected list')
                    return False
            elif multiple_fields:
                print('error, incompatible data entry')
                return False
            else:
                multiple_fields = False

        file = open(self.path, 'w')
        file.write(self.file_str)

        if multiple_fields:
            for index in range(entries):
                entry = []
                for field in self.headers:
                    if field in in_dict:
                        entry.append(in_dict[field][index])
                    else:
                        entry.append(None)
                file.write(arr_to_entry(entry))
        else:
            entry = []
            for field in self.headers:
                if field in in_dict:
                    entry.append(in_dict[field])
                else:
                    entry.append(None)
            file.write(arr_to_entry(entry))

        file.close()
        return True


if __name__ == "__main__":

    test = ObjectifyXL('../src_files/xl.xlsx')

    keys = [1, 2, 3]
    new_data = {
        'C1': ['x', 'y', 'z'],
        'C2': ['ing', 'hope', 'x'],
        'C3': ['this', 'it', 'y']
    }
    test.write_to_file(new_data, sheet_name='Sheet1', keys=keys)

    # QA = test.get_by_field({'F2': 'a'})
    # print("QA: " + str(QA))
    for each, thing in test.data.items():
        print(each, thing)
